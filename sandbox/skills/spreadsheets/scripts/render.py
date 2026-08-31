from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INVALID_SHEET_TITLE = re.compile(r"[\\/*?:\[\]]")


def _input() -> list[dict]:
    value = json.load(sys.stdin)
    workbook = value.get("workbook") if isinstance(value, dict) else None
    sheets = workbook.get("sheets") if isinstance(workbook, dict) else None
    if not isinstance(sheets, list) or not 1 <= len(sheets) <= 16:
        raise ValueError("workbook.sheets must contain 1 to 16 sheets")
    names: set[str] = set()
    for sheet in sheets:
        if not isinstance(sheet, dict):
            raise ValueError("each sheet must be an object")
        name = sheet.get("name")
        rows = sheet.get("rows")
        if (
            not isinstance(name, str)
            or not name.strip()
            or len(name) > 31
            or INVALID_SHEET_TITLE.search(name)
            or name in names
        ):
            raise ValueError("sheet name is invalid or duplicated")
        if not isinstance(rows, list) or not rows or len(rows) > 2_000:
            raise ValueError("sheet rows must contain 1 to 2000 rows")
        if any(not isinstance(row, list) or len(row) > 64 for row in rows):
            raise ValueError("each workbook row must contain at most 64 cells")
        if any(
            value is not None
            and not isinstance(value, (str, int, float, bool))
            for row in rows
            for value in row
        ):
            raise ValueError("workbook cells must be JSON scalar values")
        normalized_name = name.casefold()
        if normalized_name in names:
            raise ValueError("sheet name is invalid or duplicated")
        names.add(normalized_name)
    return sheets


def _width(value: object) -> int:
    text = "" if value is None else str(value)
    wide = sum(2 if ord(character) > 127 else 1 for character in text)
    return min(max(wide + 2, 10), 40)


def _wrapped_lines(value: object, column_width: float) -> int:
    text = "" if value is None else str(value)
    if not text:
        return 1
    capacity = max(int(column_width) - 2, 1)
    return sum(
        max(
            1,
            (
                sum(2 if ord(char) > 127 else 1 for char in line)
                + capacity
                - 1
            )
            // capacity,
        )
        for line in text.splitlines() or [""]
    )


def _horizontal_alignment(value: object) -> str:
    if isinstance(value, bool):
        return "center"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "right"
    if isinstance(value, str) and value.startswith("="):
        return "right"
    return "left"


def _verify_workbook(workbook, sheets: list[dict]) -> None:
    formula_errors = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    for sheet_spec in sheets:
        sheet = workbook[sheet_spec["name"]]
        rows = sheet_spec["rows"]
        if sheet.max_row < len(rows):
            raise ValueError("generated XLSX lost workbook rows")
        for row_index, row in enumerate(rows, start=1):
            for column_index, expected in enumerate(row, start=1):
                actual = sheet.cell(row_index, column_index).value
                if actual != expected:
                    raise ValueError("generated XLSX changed a workbook value")
                if isinstance(actual, str) and any(error in actual for error in formula_errors):
                    raise ValueError("generated XLSX contains a formula error")
        if sheet["A1"].alignment.horizontal != "center":
            raise ValueError("generated XLSX header alignment is not centered")


def _render(sheets: list[dict], output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternate_fill = PatternFill("solid", fgColor="F4F7FB")
    header_font = Font(color="FFFFFF", bold=True)
    row_border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for sheet_spec in sheets:
        sheet = workbook.create_sheet(sheet_spec["name"])
        rows = sheet_spec["rows"]
        for row in rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = sheet_spec.get("freeze_panes") or (
            "A2" if len(rows) > 1 else None
        )
        if sheet_spec.get("auto_filter", True) and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
        for column_index in range(1, sheet.max_column + 1):
            width = max(
                _width(sheet.cell(row_index, column_index).value)
                for row_index in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        column_widths = [
            sheet.column_dimensions[get_column_letter(index)].width or 10
            for index in range(1, sheet.max_column + 1)
        ]
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            row_lines = 1
            for column_index, cell in enumerate(row):
                value = cell.value
                if row_index > 1 and row_index % 2 == 0:
                    cell.fill = alternate_fill
                cell.border = row_border
                cell.alignment = Alignment(
                    horizontal=(
                        "center"
                        if cell.row == 1
                        else _horizontal_alignment(value)
                    ),
                    vertical="center" if cell.row == 1 else "top",
                    wrap_text=True,
                )
                if value is not None:
                    row_lines = max(
                        row_lines,
                        _wrapped_lines(value, column_widths[column_index]),
                    )
            if row_index > 1:
                sheet.row_dimensions[row_index].height = max(20, 16 * row_lines)
    workbook.active = 0
    workbook.save(output_path)


def main() -> None:
    output_path = Path(os.environ["NEXAFLOW_OUTPUT_PATH"])
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError("spreadsheets Skill requires a .xlsx filename")
    sheets = _input()
    _render(sheets, output_path)
    verified = load_workbook(output_path, read_only=False, data_only=False)
    if verified.sheetnames != [sheet["name"] for sheet in sheets]:
        raise ValueError("generated XLSX sheet names do not match the request")
    if any(verified[name].max_row < 1 for name in verified.sheetnames):
        raise ValueError("generated XLSX contains an empty sheet")
    _verify_workbook(verified, sheets)
    print(
        json.dumps(
            {
                "renderer": "spreadsheets",
                "sheets": len(verified.sheetnames),
                "rows": sum(verified[name].max_row for name in verified.sheetnames),
            },
            separators=(",", ":"),
        )
    )


if __name__ == "__main__":
    main()
