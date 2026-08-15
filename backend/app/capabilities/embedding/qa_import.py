import csv
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.capabilities.embedding.pipeline import KnowledgePipelineError

MAX_QA_ROWS = 5_000
MAX_QA_QUESTION_CHARS = 2_000
MAX_QA_ANSWER_CHARS = 20_000
QA_HEADER_ALIASES = {
    "question": {"question", "问题"},
    "answer": {"answer", "答案"},
    "source": {"source", "来源"},
}


@dataclass(frozen=True)
class QaRow:
    question: str
    answer: str
    source: str
    row_number: int


def _cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def _semantic_header(value: object) -> str | None:
    header = _cell(value)
    for semantic, aliases in QA_HEADER_ALIASES.items():
        if header in aliases or header.lower() in aliases:
            return semantic
    return None


def validate_qa_rows(rows: Iterable[Sequence[object]]) -> list[QaRow]:
    iterator = iter(enumerate(rows, start=1))
    header_indexes: dict[str, int] | None = None
    for row_number, row in iterator:
        if not any(_cell(value) for value in row):
            continue
        header_indexes = {}
        for index, value in enumerate(row):
            semantic = _semantic_header(value)
            if semantic is None:
                continue
            if semantic in header_indexes:
                raise KnowledgePipelineError(
                    f"QA import row {row_number} has duplicate {semantic} headers."
                )
            header_indexes[semantic] = index
        missing = {"question", "answer"} - set(header_indexes)
        if missing:
            raise KnowledgePipelineError(
                "QA import requires question/问题 and answer/答案 headers."
            )
        break
    if header_indexes is None:
        raise KnowledgePipelineError("QA import has no header row.")

    drafts: list[QaRow] = []
    for row_number, row in iterator:
        values = [_cell(value) for value in row]
        if not any(values):
            continue

        def value_for(field: str) -> str:
            index = header_indexes.get(field)
            return values[index] if index is not None and index < len(values) else ""

        question = value_for("question")
        answer = value_for("answer")
        if not question:
            raise KnowledgePipelineError(
                f"QA import row {row_number} has an empty question."
            )
        if not answer:
            raise KnowledgePipelineError(
                f"QA import row {row_number} has an empty answer."
            )
        if len(question) > MAX_QA_QUESTION_CHARS:
            raise KnowledgePipelineError(
                f"QA import row {row_number} question exceeds {MAX_QA_QUESTION_CHARS} characters."
            )
        if len(answer) > MAX_QA_ANSWER_CHARS:
            raise KnowledgePipelineError(
                f"QA import row {row_number} answer exceeds {MAX_QA_ANSWER_CHARS} characters."
            )
        if len(drafts) >= MAX_QA_ROWS:
            raise KnowledgePipelineError(
                f"QA import row {row_number} exceeds the {MAX_QA_ROWS}-row limit."
            )
        drafts.append(
            QaRow(
                question=question,
                answer=answer,
                source=value_for("source"),
                row_number=row_number,
            )
        )
    if not drafts:
        raise KnowledgePipelineError("QA import has no data rows.")
    return drafts


def extract_qa_rows(filename: str, path: Path) -> list[QaRow]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                return validate_qa_rows(csv.reader(stream))
        except UnicodeDecodeError as exc:
            raise KnowledgePipelineError("QA CSV must use UTF-8 encoding.") from exc
    if suffix == ".xlsx":
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                return validate_qa_rows(
                    workbook.active.iter_rows(values_only=True)
                )
            finally:
                workbook.close()
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise KnowledgePipelineError("QA XLSX file is invalid.") from exc
    raise KnowledgePipelineError("QA import supports CSV and XLSX files only.")
